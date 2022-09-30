import datetime
import openpyxl

class Writer(object):
    def write_shift(self, is_nightshift, today_shift, report_filename, w_hours, core_time, incoming_time, outgoing_time, work_total):
        today = datetime.date.today()
        def func(_):
            if _ == '公' or _ == '法':
                return '公法'
            elif _ == '有':
                return '事前有給'
            else:
                return _

        today_shift = func(today_shift)

        wb = openpyxl.load_workbook(report_filename)
        ws = wb.worksheets[0]

        d = int(today.strftime('%d'))
        if is_nightshift:
            t = 7
        else:
            t = 8

        c_shift = ws['D{}'.format(str(d + t))]
        c_w_hours = ws['J{}'.format(str(d + t))]
        c_core_time = ws['L{}'.format(str(d + t))]
        c_inc_time = ws['BG{}'.format(str(d + t))]
        c_ougo_time = ws['BH{}'.format(str(d + t))]
        c_work_total = ws['BI{}'.format(str(d + t))]

        c_shift.value = today_shift
        if today_shift != '公法' and today_shift != '事前有給':
            c_w_hours.value = w_hours
            c_core_time.value = core_time
            c_inc_time.value = incoming_time
            c_ougo_time.value = outgoing_time
            c_work_total.value = work_total

        wb.save(report_filename)


